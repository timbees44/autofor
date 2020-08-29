import os
import openpyxl
import csv
import binascii
import subprocess
from datetime import datetime
import hashlib
import pandas as pd
from simple_term_menu import TerminalMenu
from tkinter import *
from tkinter import filedialog
import re
from importlib import resources

# a tool class to store the forensic methods we might want to use


class Tools:

    def __init__(self):
        pass

    # find fullpath of files required by tools that come with programme
    def fullpath(self, file):
        return os.path.dirname(
            os.path.abspath(file))

    # method to identify files in evidence
    def files(self, sspath, mount, case):
        print("starting files function...")
        log_path = f"{sspath}/logical"
        """
        Check if file is in securestore logical directory or a mount directory
        This will initially be the only two places where evidence will be available
        to examine
        """
        if mount == "":
            path = log_path
        else:
            path = mount
        # open worksheet and set to file_info sheet
        workbook = f"{sspath}/analysis/spreadsheet_{case}.xlsx"
        wb = openpyxl.load_workbook(workbook)
        ws = wb["file_info"]
        # get signature csv
        sigcsv = sigcsv = os.path.dirname(
            os.path.abspath("config_files/file_sigs.csv") + "/file_sig.csv")
        # iterate through all files and sub file
        count = 2
        for root, dirs, files in os.walk(path, topdown=False):
            for name in files:

                # add full paths to first column
                p = os.path.join(root, name)
                ws["A" + str(count)] = p
                # strip extension
                p_ext = str(p.split(".")[-1])
                ws["B" + str(count)] = p_ext

                # open p (full path) and convert to hex)
                fsigs = pd.read_csv(sigcsv)
                # open p (full path) and convert to hex
                with open(p, 'rb') as f:
                    content = f.read()
                    hexed = binascii.hexlify(content).decode("utf-8")
                    # iterate through columns

                    for index, row in fsigs.iterrows():
                        # check if file_sig at correct offset
                        head = str(row["header"]).replace(" ", "")
                        if hexed.upper().startswith(head):
                            sig_len = len(row["header"])
                            ws["C" + str(count)] = row["extension"]
                            ws["D" + str(count)] = row["header"]
                            ws["E" + str(count)] = row["offset"]
                            ws["F" + str(count)] = row["description"]
                            # check for longest matching signature
                            if sig_len > sig_len:
                                ws["C" + str(count)] = row["extension"]
                                ws["D" + str(count)] = row["header"]
                                ws["E" + str(count)] = row["offset"]
                                ws["F" + str(count)] = row["description"]
                            # set datetime
                            ws["H" + str(count)] = str(datetime.now())
                        elif str(row["header"]) in hexed.upper() and ws["C" + str(count)] == "":
                            ws["G" + str(count)] = row["extension"] + \
                                ":" + row["header"]
                            # add count to move through rows
                count += 1

                # add count to move through files

        wb.save(workbook)

        # exiftool output csv
        exif_out = f"{sspath}/analysis/exiftool_out_{case}.csv"
        # hidden directory csv exiftool output
        hid_exif_out = f"{sspath}/analysis/hidden_exiftool_out_{case}.csv"
        # hidden directory exiftool check
        for root, dirs, files in os.walk(path, topdown=False):
            for subdir in dirs:
                if subdir.startswith("."):
                    x = os.path.join(path, root, subdir)
                    try:
                        subprocess.check_output(
                            f"exiftool -csv -time:all -r -a -u \"{x}\" >> {hid_exif_out}", shell=True)
                    except:
                        pass
        try:
            subprocess.check_output(
                f"exiftool -csv -time:all -r \"{path}\" >> {exif_out}", shell=True)
        except:
            pass

   # intial hash function to be used when file/dir first located on system

    def inithash(self, case, sspath, evidence):
        sheetpath = f"{sspath}/analysis/spreadsheet_{case}.xlsx"
        print("hashing, please wait...")
        # call excel template
        # check if excel sheet has already been created so that it doesn't get overwritten by calling template each time
        if not os.path.isfile(sheetpath):
            wb = openpyxl.load_workbook(os.path.dirname(os.path.abspath(
                "excel_temp.xlsx")) + "/config_files/excel_temp.xlsx")
            wb.save(sheetpath)

        # if a directory is selected it will be hashed in it's current location and an inital hash list will be added to the first sheet of the spreadsheet
        wb = openpyxl.load_workbook(sheetpath)
        ws = wb["initial_hashes"]
        # hash initial directory
        if os.path.isdir(evidence):
            count = 2
            for root, dirs, files in os.walk(evidence, topdown=False):
                for name in files:
                    # count for skipping title cells and iteration if dir
                    x = os.path.join(root, name)
                    ws["A" + str(count)] = x
                    sha256 = hashlib.sha256()
                    # breaking into blocks to handle larger files
                    block_size = 4096
                    with open(x, 'rb') as f:
                        for block in iter(lambda: f.read(block_size), b''):
                            sha256.update(block)
                            ws["B" + str(count)] = sha256.hexdigest()

                    count += 1

        # hash initial file
        elif os.path.isfile(evidence):
            sha256 = hashlib.sha256()
            with open(evidence, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    sha256.update(chunk)
                    ws["A2"] = str(evidence)
                    ws["B2"] = sha256.hexdigest()

        wb.save(sheetpath)

    # hash function to be used everytime after initial hash. it will add sheets to spreadsheet
    def hash(self, case, path, sspath):
        print(f"hashing {path}")
        # set sheetpath
        sheetpath = f"{sspath}/analysis/spreadsheet_{case}.xlsx"
        print("hashing, please wait...")

        wb = openpyxl.load_workbook(sheetpath)
        # sheets = wb.sheetnames
        date_time = datetime.now()
        sheet_title = f"hashes{date_time}".replace(":", "")
        # print(sheet_title)
        wb.create_sheet(sheet_title)
        ws = wb[sheet_title]
        ws["A1"] = "path"
        ws["B1"] = "hash"
        # hash initial directory
        if os.path.isdir(path):
            x_count = 2
            for root, dirs, files in os.walk(path, topdown=False):
                for name in files:
                    x = os.path.join(root, name)
                    ws["A" + str(x_count)] = x
                    sha256 = hashlib.sha256()
                    # breaking into blocks to handle larger files
                    block_size = 4096
                    with open(x, 'rb') as f:
                        for block in iter(lambda: f.read(block_size), b''):
                            sha256.update(block)
                            ws["B" + str(x_count)] = sha256.hexdigest()
                        x_count += 1

        # hash initial file
        elif os.path.isfile(path):
            sha256 = hashlib.sha256()
            with open(path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    sha256.update(chunk)
                    ws["A2"] = path
                    ws["B2"] = sha256.hexdigest()

        wb.save(sheetpath)
        print("hash complete!")

    # build a text based directory tree representation
    def dir_tree(self, mount, sspath):
        # chose mount location for disk images or logical if not
        if mount == "":
            path = sspath + "/logical"
        else:
            path = mount

        # output of directory tree
        dir_tree = sspath + "/analysis/directory_tree.txt"
        if not os.path.isfile(dir_tree):
            os.system(f"touch {dir_tree}")
        # handle a mounted file and regular directories with subdirs
            f = open(dir_tree, "a")
            for root, dirs, files in os.walk(path):
                level = root.replace(path, '').count(os.sep)
                indent = ' ' * 4 * (level)
                f.write('\n{}{}/'.format(indent, os.path.basename(root)))
                subindent = ' ' * 4 * (level + 1)
                for i in files:
                    f.write('\n{}{}'.format(subindent, i))

            f.write("\n")
            f.close()
            print(
                f"filesystem tree succesfully written to {dir_tree}")
        else:
            pass

    # terminal based directory tree for quick reference to evidence structure
    # to be implemented at a later stage - low priority

    def dir_tree_terminal(self, mount, sspath):
        # chose mount location for disk images or logical if not
        if mount == "":
            path = sspath + "/logical"
        else:
            path = mount

        # escape codes for colours if output directly to terminal
        # different colours for files/folders
        foldb = "\033[94m"
        folde = "\033[0m"
        fileb = "\033[92m"
        filee = "\033[0m"

        for root, dirs, files in os.walk(path):
            level = root.replace(path, '').count(os.sep)
            indent = ' ' * 4 * (level)
            print('{}{}/'.format(indent, foldb + os.path.basename(root) + folde))
            subindent = ' ' * 4 * (level + 1)
            for i in files:
                print('{}{}'.format(subindent, fileb + i + filee))
            print("\n")

    """search for partiular word(s) in all files or run a dictionary attack against one file
    this will check strings output, and the physical content of the file"""

    def wordsearch(self, sspath, mount, wordlist="none"):
        word_menu_title = "Here you can search for words or phrases in evidence files/directories" + \
            f"\nWordlist in Use: {wordlist}"
        word_menu_items = ["Search specific file", "Search specific directory",
                           "Search all evidence", "Select Wordlist"]
        word_menu_exit = False

        word_menu = TerminalMenu(title=word_menu_title,
                                 menu_entries=word_menu_items)

        # set global path variable to be changed
        path = ""

        while not word_menu_exit:
            item_sel = word_menu.show()

            if item_sel == 0:
                root = Tk()
                root.withdraw()
                path = root.file = filedialog.askopenfilename()

            elif item_sel == 1:
                root = Tk()
                root.withdraw()
                path = root.directory = filedialog.askdirectory()

            elif item_sel == 2:
                if sspath == "":
                    path = mount
                else:
                    path = sspath

            elif item_sel == 3:
                root = Tk()
                root.withdraw()
                wordlist = root.file = filedialog.askopenfilename()
                os.system("clear")
                self.wordsearch(sspath, mount, wordlist=wordlist)
                return

            word_menu_exit = True

        # create empty wordlist
        w_list = []
        # check if wordlist selected
        if os.path.isfile(wordlist):
            with open(wordlist, "r") as w:
                for i in w:
                    w_list.append(i)

        else:
            print("Please input word(s) separated by commas")
            words = input("Note that commas will be used as separators: ")
            w_list = result = [w.strip() for w in words.split(',')]

        print(f"Here is your word list:\n{w_list}")
        print(f"{path} will be searched for matching strings")
        print("Checking file(s) now...")

        # iteration count for no results
        count = 0

        # check if file or dir
        if os.path.isdir(path):
            for root, dirs, files in os.walk(path):
                for file in files:
                    p = os.path.join(root, file)
                    f = subprocess.check_output(["strings", p]).decode("utf-8")
                    for word in w_list:
                        if re.findall(word, f):
                            print(word + " found in: " + p)
                        else:
                            count += 1

        elif os.path.isfile(path):
            f = subprocess.check_output(["strings", path]).decode("utf-8")
            for word in w_list:
                if re.findall(word, f):
                    print(word + " found in " + [path])
                else:
                    count += 1

        if count > 0:
            print(f"No matching strings found in {path}...")

    def regexsearch(self, sspath, mount, case):
        # may be too lenient
        # modified from emailregex.com
        re_email = re.compile(r"(^[\w.+-]+@[\w-]+\.[a-zA-Z0-9-.])+$")

        # the following taken from urlregex.com, ipregex.com and phoneregex.com

        # requires start with http(s) to be recognised. Might miss some urls but do remove this condition allows of noise
        re_url = re.compile(
            r"http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+")

        # ip addresses
        re_ip = re.compile(
            r"^(([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])\.){3}([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])$")

        # UK phone numbers
        re_phone = re.compile(
            r"^(?:(?:\(?(?:0(?:0|11)\)?[\s-]?\(?|\+)44\)?[\s-]?(?:\(?0\)?[\s-]?)?)|(?:\(?0))(?:(?:\d{5}\)?[\s-]?\d{4,5})|(?:\d{4}\)?[\s-]?(?:\d{5}|\d{3}[\s-]?\d{3}))|(?:\d{3}\)?[\s-]?\d{3}[\s-]?\d{3,4})|(?:\d{2}\)?[\s-]?\d{4}[\s-]?\d{4}))(?:[\s-]?(?:x|ext\.?|\#)\d{3,4})?$")

        # check if disk image or not
        if mount == "":
            path = sspath + "/logical"
        else:
            path = mount

        # open workbook
        workbook = f'{sspath}/analysis/spreadsheet_{case}.xlsx'
        wb = openpyxl.load_workbook(workbook)
        ws = wb["regex"]

        # excel count
        count = 2
        # iterate through evidence
        for root, dirs, files in os.walk(path):
            for file in files:
                print(file)
                p = os.path.join(root, file)
                # set strings variable
                f = subprocess.check_output(["strings", p]).decode("utf-8")
                # print(f)
                list_f = (f.split(" "))
                # print(list_f)
                # more concise regex variables to allow for easy printing
                for i in list_f:
                    """
                    Using re.search to find individual matches for each line.
                    Lines have been simulated using a list created with "\r"
                    characters that are present in the strings output.
                    Strings was chosen over opening the file as many files 
                    can't be open and read.
                    Re.search, may not be as effective as re.iterate as it may
                    miss items on some lines. However at the moment it seems to
                    work the best with the spreadsheet setup.
                    """
                    if re_email.search(str(i)):
                        print(i)
                        ws["A" + str(count)] = p
                        ws["B" + str(count)] = p.split("/")[-1]
                        ws["E" + str(count)] = str(i)
                        ws["G" + str(count)] = list_f.index(i) + 1
                        count += 1

                    elif re_ip.search(i):
                        print(i)
                        ws["A" + str(count)] = p
                        ws["B" + str(count)] = p.split("/")[-1]
                        ws["C" + str(count)] = i
                        ws["G" + str(count)] = list_f.index(i) + 1
                        count += 1

                    elif re_phone.search(i):
                        print(i)
                        ws["A" + str(count)] = p
                        ws["B" + str(count)] = p.split("/")[-1]
                        ws["D" + str(count)] = i
                        ws["G" + str(count)] = list_f.index(i) + 1
                        count += 1

                    elif re_url.search(i):
                        print(i)
                        ws["A" + str(count)] = p
                        ws["B" + str(count)] = p.split("/")[-1]
                        ws["F" + str(count)] = i
                        ws["G" + str(count)] = list_f.index(i) + 1
                        count += 1

        wb.save(workbook)

    def carver(self, sspath, mount, case):
        print("Carving files to check for embedded/composite files...")
        # check all files in mount location or sercure store
        if mount == "":
            path = sspath + "/logical"
        else:
            path = mount

        if "foremost" not in os.listdir(sspath + "/tests"):
            os.system(f"mkdir {sspath}/tests/foremost")

        # open workbook to add carved files
        workbook = f'{sspath}/analysis/spreadsheet_{case}.xlsx'
        wb = openpyxl.load_workbook(workbook)
        ws = wb["carved_files"]
        count = 2
        # iterate through all files in evidence
        for root, dirs, files in os.walk(path):
            for file in files:
                # set full path
                p = os.path.join(root, file)
                # add path to worksheet
                ws["A" + str(count)] = p
                # add filename to worksheet
                ws["B" + str(count)] = p.split("/")[-1]
                # set file friendly datetime.now() variable to maintain chaing of analysis
                dt = str(datetime.now()).replace(":", "-").replace(" ", ".")
                # add datetime to worksheet
                ws["D" + str(count)] = str(datetime.now())
                # create unique foremost ouput dirs
                out_dir = sspath + f"/tests/foremost/fm_{file}_{dt}"
                # add output path to excel
                ws["C" + str(count)] = out_dir
                # run foremost against each file with verbose specifying that all file types are to be retrieved
                subprocess.check_output(
                    f"foremost -Q -t all '{p}' -o '{out_dir}'", shell=True)
                # if no files are extracted, there will only be "audit.txt" in output directory
                if len(os.listdir(out_dir)) == 1:
                    # automatically delete directories with no extractions as not useful for this part of investigation
                    os.system(f"rm -r '{out_dir}'")

                # add to count
                count += 1

        wb.save(workbook)
